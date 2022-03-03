import openpyxl
from PIL import Image, UnidentifiedImageError
import os

path = "/Users/arturumagulov/Documents/Suvar/site/wwwroot"


def write_xls(lst, to_path):
    """ write_xls(result, "result.xlsx") """

    wb = openpyxl.Workbook()
    wb.create_sheet(title="Лист", index=0)
    sheet = wb["Лист"]
    for i in range(len(lst)):
        for j in lst[i]:
            value = str(j)
            cell = sheet.cell(row=i + 1, column=(lst[i].index(j)) + 1)
            cell.value = value
    wb.save(to_path)


def conv_to_webp(path, name):

    sub_result = list()

    clear_path = f"{path}/{name}"
    sub_result.append(clear_path[42:])
    sub_result.append(round(os.stat(clear_path).st_size / (1024 * 1024), 2))
    im = Image.open(clear_path).convert("RGB")
    sub_result.append("конвертированно")
    im.save(f"{clear_path[:-3]}webp", "webp")
    sub_result.append(f"{clear_path[42:-3]}webp")
    sub_result.append(round(os.stat(f"{clear_path[:-3]}webp").st_size / (1024 * 1024), 2))
    os.remove(clear_path)

    return sub_result


def path_remove(path):
    result = []
    for path, dirs, files in os.walk(path):
        for file in files:
            if file[-3:] == "jpg":
                result.append([path, file])

            elif file[-3:] == 'png':
                result.append([path, file])

            elif file[-3:] == 'JPG':
                result.append([path, file])

            # elif file[-4:] == 'jpeg':
            #     result.append([path, file])

    return result


if __name__ == '__main__':
    data = path_remove(path)
    logs = list()
    error_logs = list()
    for i in data:
        clear_path = f"{i[0]}/{i[1]}"
        try:
            logs.append(conv_to_webp(i[0], i[1]))
            print("logs", logs)

        except UnidentifiedImageError:
            print(f"{clear_path[42:]} ошибка")
            error_logs.append([clear_path[42:], "ошибка"])
        except FileNotFoundError:
            print(f"{clear_path[42:]} файл не найден")
            error_logs.append([clear_path[42:], "файл не найден"])
        except ValueError:
            print(f"{clear_path[42:]} большой размер данных")
            error_logs.append([clear_path[42:], "большой размер данных"])
    write_xls(logs, "logs.xlsx")
    write_xls(error_logs, "errors.xlsx")

